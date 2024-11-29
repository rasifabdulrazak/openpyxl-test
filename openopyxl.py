
from openpyxl import Workbook
from openpyxl.styles import NamedStyle
from datetime import datetime,date,time as time_dt


workbook = Workbook()
sheet = workbook.active

# Add a date and apply a date format
date_style = NamedStyle(name="date_style", number_format='dd-mm-yyyy')
time_style = NamedStyle(name="time_style", number_format="[$-F400]h:mm AM/PM")
datetime_style = NamedStyle(name="datetime_style", number_format='dd-mm-yyyy hh:mm')
workbook.add_named_style(date_style)
workbook.add_named_style(time_style)
workbook.add_named_style(datetime_style)

print(datetime(2024,7,1))

# Add a date, time, and datetime to the sheet
sheet["A1"] = "Date"
sheet["A2"] = datetime(2024, 7, 1)  # Use your desired date
sheet["A2"].style = date_style

sheet["B1"] = "Time"
sheet["B2"] = time_dt(hour=14, minute=30)  # 14:30 (2:30 PM)
sheet["B2"].style = time_style

sheet["C1"] = "DateTime"
sheet["C2"] = datetime(2024, 7, 1, 14, 30)  # Use your desired datetime
sheet["C2"].style = datetime_style


# Save the workbook
# workbook.save(filename="formatted_cells.xlsx")

workbook.save(filename="htest.xlsx")

# dictoi = {'a':1,'b':2,'c':1,'d':2,'e':3}
# # dictoi.pop('a')
# # dictoi['f'] = 6
# # dictoi.a
# print(dictoi)
# result = {1:['a','c'],2:['b','d'],3:['e']}

# res = {}
# for key,value in dictoi.items():
#     print(value,key,"===")
#     if value not in res:
#         res[value] = []
#     res[value].append(key)

# print(res)
# defaults = dict.fromkeys(["name", "age", "city"], "Unknown")
# print(defaults)

# st = "Rasif Rasif"
# f = {}

# for i in st:
#     f[i] = f.get(i,0) + 1

# print(f)

# d = {'name':'ras','age':90}
# # d['city'] = 'kolampur'
# # # del d['age']
# print(d.items())
# for i,j in d.items():
#     print(i,j)
# d = dictoi | result
# print(d)

# sq = {i:i**2 for i in range(1,10)}
# print(sq)


# filter_list.py

# MILLION_NUMBERS = list(range(1_000_000))

# def for_loop():
#     output = []
#     for element in MILLION_NUMBERS:
#         if not element % 2:
#             output.append(element)
#     return output

# MILLION_NUMBERS = list(range(1_000_000))

# def list_comprehension():
#     return [number for number in MILLION_NUMBERS if not number % 2]

# tu = tuple(1,2,3,3)
# print(tu)

# for i,j in enumerate(d):
#     print(i,j)

# questions = ['name', 'location', 'favorite language']
# answers = ['Codersdaily', 'Indore', 'Python']
# for q, a in zip(questions, answers):
#     print('What is your {0}?  It is {1}.'.format(q, a))


# import xlsxwriter
# from datetime import datetime, date, time as time_dt

# # Create a new Excel file and add a worksheet.
# workbook = xlsxwriter.Workbook('demo_with_datetime.xlsx',{
#             'date_format': 'dd-mm-yyyy',
#             'datetime_format': 'dd-mm-yyyy hh:mm',
#             'time_format': 'hh:mm'
#         })
# worksheet = workbook.add_worksheet()

# # Widen the first column to make the text clearer.
# worksheet.set_column('A:A', 20)

# # Add a bold format to use to highlight cells.
# bold = workbook.add_format({'bold': True})

# # Write some simple text.
# worksheet.write('A1', 'Hello')

# # Text with formatting.
# worksheet.write('A2', 'World', bold)

# # Write some numbers, with row/column notation.
# worksheet.write(2, 0, 123)
# worksheet.write(3, 0, 123.456)

# # Define datetime, date, and time formats
# datetime_format = workbook.add_format({'num_format': 'dd-mm-yyyy hh:mm'})
# date_format = workbook.add_format({'num_format': 'dd-mm-yyyy'})
# time_format = workbook.add_format({'num_format': 'hh:mm'})

# # Write datetime, date, and time values
# worksheet.write_datetime('A4', datetime(2024, 7, 1, 10, 30), datetime_format)
# worksheet.write_datetime('A5', date(2024, 7, 1), date_format)
# worksheet.write('A6', time_dt(10, 30).strftime('%H:%M'), time_format)

# # Insert an image.
# # worksheet.insert_image('B8', 'logo.png')

# # Close the workbook to save the file.
# workbook.close()

# print("File saved successfully as 'demo_with_datetime.xlsx'")

# names_list = ["Alice", "Bob", "Alice", "Charlie", "Alice", "Bob"]

# g = {}

# for i in names_list:
#     if i in g:
#         g[i] += 1
#         break
#     else:
#         g[i] = 1

# print(g)

# import random

# fruit=['apple', 'banana', 'papaya', 'cherry']

# random.shuffle(fruit)
# print(fruit)

# from datetime import datetime
# import xlsxwriter

# # Create a workbook and add a worksheet.
# workbook = xlsxwriter.Workbook("datetimes.xlsx")
# worksheet = workbook.add_worksheet()
# bold = workbook.add_format({"bold": True})

# # Expand the first columns so that the dates are visible.
# worksheet.set_column("A:B", 30)

# # Write the column headers.
# worksheet.write("A1", "Formatted date", bold)
# worksheet.write("B1", "Format", bold)

# # Create a datetime object to use in the examples.

# date_time = datetime.strptime("2013-01-23 12:30:05.123", "%Y-%m-%d %H:%M:%S.%f")

# # Examples date and time formats. In the output file compare how changing
# # the format codes change the appearance of the date.
# date_formats = (
#     "dd-mm-yyyy",
#     "dd-mm-yy",
#     "dd/mm/yy",
#     "mm/dd/yy",
#     "dd m yy",
#     "d mm yy",
#     "d mmm yy",
#     "d mmmm yy",
#     "d mmmm yyy",
#     "d mmmm yyyy",
#     "dd/mm/yy hh:mm",
#     "dd/mm/yy hh:mm:ss",
#     "dd/mm/yy hh:mm:ss.000",
#     "hh:mm",
#     "hh:mm:ss",
#     "hh:mm:ss.000",
# )

# # Start from first row after headers.
# row = 1

# # Write the same date and time using each of the above formats.
# for date_format_str in date_formats:
#     # Create a format for the date or time.
#     date_format = workbook.add_format({"num_format": date_format_str, "align": "left"})

#     # Write the same date using different formats.
#     worksheet.write_datetime(row, 0, date_time, date_format)

#     # Also write the format string for comparison.
#     worksheet.write_string(row, 1, date_format_str)

#     row += 1

# workbook.close()